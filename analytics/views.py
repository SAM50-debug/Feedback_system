from django.shortcuts import render, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count
from django.http import HttpResponse
from forms_app.models import FeedbackForm, Question, Response, MCQOption
from core.models import School, Department, Course
from accounts.models import Student, StudentCourse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

@staff_member_required
def analytics_dashboard(request):
    schools = School.objects.all()
    selected_school = request.GET.get('school')
    selected_department = request.GET.get('department')
    selected_course = request.GET.get('course')
    
    departments = Department.objects.none()
    courses = Course.objects.none()
    forms = FeedbackForm.objects.none()
    
    if selected_school:
        departments = Department.objects.filter(school_id=selected_school)
    
    if selected_department:
        courses = Course.objects.filter(department_id=selected_department)
    
    if selected_course:
        forms = FeedbackForm.objects.filter(course_id=selected_course).select_related(
            'course', 'teacher', 'course__department', 'course__department__school'
        )
    
    context = {
        'schools': schools,
        'departments': departments,
        'courses': courses,
        'forms': forms,
        'selected_school': selected_school,
        'selected_department': selected_department,
        'selected_course': selected_course,
    }
    return render(request, 'analytics/dashboard.html', context)

@staff_member_required
def form_results(request, form_id):
    form = get_object_or_404(
        FeedbackForm.objects.select_related(
            'course', 'teacher', 'course__department', 'course__department__school'
        ), 
        id=form_id
    )
    
    questions = form.questions.all().prefetch_related('options')
    total_submissions = form.submissions.count()
    
    results = []
    
    for question in questions:
        question_data = {
            'question': question,
            'total_responses': 0,
            'data': []
        }
        
        if question.question_type == 'mcq':
            # Get all options for this question first
            all_options = question.options.all().order_by('order')
            
            # Get response counts for each option
            option_counts = Response.objects.filter(
                question=question,
                mcq_answer__isnull=False
            ).values(
                'mcq_answer__id',
                'mcq_answer__option_text',
                'mcq_answer__order'
            ).annotate(
                count=Count('id')
            ).order_by('mcq_answer__order')
            
            # Convert to dict for easier lookup
            counts_dict = {item['mcq_answer__id']: item for item in option_counts}
            
            # Build complete data including options with 0 responses
            complete_data = []
            for option in all_options:
                if option.id in counts_dict:
                    complete_data.append(counts_dict[option.id])
                else:
                    # Add option with 0 count
                    complete_data.append({
                        'mcq_answer__id': option.id,
                        'mcq_answer__option_text': option.option_text,
                        'mcq_answer__order': option.order,
                        'count': 0
                    })
            
            question_data['total_responses'] = sum(item['count'] for item in complete_data)
            question_data['data'] = complete_data
            
        elif question.question_type == 'text':
            # Get all text responses with student info
            text_responses = Response.objects.filter(
                question=question,
                text_answer__isnull=False
            ).exclude(
                text_answer=''
            ).select_related(
                'submission__student'
            ).values_list(
                'text_answer',
                'submission__student__name',
                'submission__submitted_at'
            ).order_by('-submission__submitted_at')
            
            question_data['total_responses'] = text_responses.count()
            question_data['data'] = list(text_responses)
        
        results.append(question_data)
    
    context = {
        'form': form,
        'results': results,
        'total_submissions': total_submissions,
    }
    return render(request, 'analytics/form_results.html', context)

@staff_member_required
def export_form_results(request, form_id):
    """Export form results to Excel file"""
    form = get_object_or_404(
        FeedbackForm.objects.select_related(
            'course', 'teacher', 'course__department', 'course__department__school'
        ), 
        id=form_id
    )
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary information
    ws_summary['A1'] = "Feedback Report"
    ws_summary['A1'].font = Font(bold=True, size=16, color="6366F1")
    
    ws_summary['A3'] = "Form Title:"
    ws_summary['B3'] = form.title
    ws_summary['A4'] = "Course:"
    ws_summary['B4'] = f"{form.course.code} - {form.course.name}"
    ws_summary['A5'] = "Teacher:"
    ws_summary['B5'] = form.teacher.name
    ws_summary['A6'] = "Department:"
    ws_summary['B6'] = form.course.department.name
    ws_summary['A7'] = "School:"
    ws_summary['B7'] = form.course.department.school.name
    ws_summary['A8'] = "Total Submissions:"
    ws_summary['B8'] = form.submissions.count()
    ws_summary['A9'] = "Generated On:"
    ws_summary['B9'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Make labels bold
    for row in range(3, 10):
        ws_summary[f'A{row}'].font = Font(bold=True)
    
    # MCQ Results Sheet
    ws_mcq = wb.create_sheet("MCQ Results")
    
    # Headers
    ws_mcq['A1'] = "Question"
    ws_mcq['B1'] = "Option"
    ws_mcq['C1'] = "Count"
    ws_mcq['D1'] = "Percentage"
    
    for col in ['A', 'B', 'C', 'D']:
        ws_mcq[f'{col}1'].fill = header_fill
        ws_mcq[f'{col}1'].font = header_font
        ws_mcq[f'{col}1'].border = border
        ws_mcq[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 2
    questions = form.questions.filter(question_type='mcq').prefetch_related('options')
    
    for question in questions:
        # Get response counts
        option_counts = Response.objects.filter(
            question=question,
            mcq_answer__isnull=False
        ).values(
            'mcq_answer__option_text'
        ).annotate(
            count=Count('id')
        ).order_by('mcq_answer__order')
        
        total_responses = sum(item['count'] for item in option_counts)
        
        # Write question and options
        first_row = current_row
        for idx, item in enumerate(option_counts):
            if idx == 0:
                ws_mcq[f'A{current_row}'] = f"Q{question.order}: {question.question_text}"
                ws_mcq[f'A{current_row}'].font = Font(bold=True)
            
            ws_mcq[f'B{current_row}'] = item['mcq_answer__option_text']
            ws_mcq[f'C{current_row}'] = item['count']
            
            if total_responses > 0:
                percentage = (item['count'] / total_responses) * 100
                ws_mcq[f'D{current_row}'] = f"{percentage:.1f}%"
            else:
                ws_mcq[f'D{current_row}'] = "0%"
            
            # Apply borders
            for col in ['A', 'B', 'C', 'D']:
                ws_mcq[f'{col}{current_row}'].border = border
            
            current_row += 1
        
        # Merge question cells
        if len(option_counts) > 1:
            ws_mcq.merge_cells(f'A{first_row}:A{current_row-1}')
        
        current_row += 1  # Empty row between questions
    
    # Text Responses Sheet
    ws_text = wb.create_sheet("Text Responses")
    
    # Headers
    ws_text['A1'] = "Question"
    ws_text['B1'] = "Student"
    ws_text['C1'] = "Response"
    ws_text['D1'] = "Submitted On"
    
    for col in ['A', 'B', 'C', 'D']:
        ws_text[f'{col}1'].fill = header_fill
        ws_text[f'{col}1'].font = header_font
        ws_text[f'{col}1'].border = border
        ws_text[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 2
    text_questions = form.questions.filter(question_type='text')
    
    for question in text_questions:
        responses = Response.objects.filter(
            question=question,
            text_answer__isnull=False
        ).exclude(
            text_answer=''
        ).select_related(
            'submission__student'
        ).order_by('submission__submitted_at')
        
        for response in responses:
            ws_text[f'A{current_row}'] = f"Q{question.order}: {question.question_text}"
            ws_text[f'B{current_row}'] = response.submission.student.name
            ws_text[f'C{current_row}'] = response.text_answer
            ws_text[f'D{current_row}'] = response.submission.submitted_at.strftime("%Y-%m-%d %H:%M")
            
            # Apply borders
            for col in ['A', 'B', 'C', 'D']:
                ws_text[f'{col}{current_row}'].border = border
            
            # Wrap text for response column
            ws_text[f'C{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            current_row += 1
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 50
    
    ws_mcq.column_dimensions['A'].width = 50
    ws_mcq.column_dimensions['B'].width = 40
    ws_mcq.column_dimensions['C'].width = 12
    ws_mcq.column_dimensions['D'].width = 12
    
    ws_text.column_dimensions['A'].width = 50
    ws_text.column_dimensions['B'].width = 25
    ws_text.column_dimensions['C'].width = 60
    ws_text.column_dimensions['D'].width = 18
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Feedback_{form.course.code}_{form.teacher.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@staff_member_required
def export_students_list(request):
    """Export all registered students to Excel file"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registered Students"
    
    # Styling
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['S.No', 'Roll Number', 'Name', 'School', 'Department', 'Enrolled Courses', 'Registration Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get all students with related data
    students = Student.objects.filter(is_staff=False).select_related(
        'school', 'department'
    ).prefetch_related('enrolled_courses__course').order_by('school', 'department', 'name')
    
    # Write data
    row_num = 2
    for idx, student in enumerate(students, 1):
        # Get enrolled courses
        courses = student.enrolled_courses.all()
        course_list = ', '.join([f"{sc.course.code}" for sc in courses]) if courses.exists() else 'None'
        
        # Write row
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=student.roll_number)
        ws.cell(row=row_num, column=3, value=student.name)
        ws.cell(row=row_num, column=4, value=student.school.name if student.school else 'N/A')
        ws.cell(row=row_num, column=5, value=student.department.name if student.department else 'N/A')
        ws.cell(row=row_num, column=6, value=course_list)
        ws.cell(row=row_num, column=7, value=student.date_joined.strftime("%Y-%m-%d"))
        
        # Apply borders
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).border = border
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 18
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "Student Registration Summary"
    ws_summary['A1'].font = Font(bold=True, size=16, color="6366F1")
    
    ws_summary['A3'] = "Total Students:"
    ws_summary['B3'] = students.count()
    ws_summary['A4'] = "Generated On:"
    ws_summary['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # School-wise count
    ws_summary['A6'] = "School-wise Distribution:"
    ws_summary['A6'].font = Font(bold=True)
    row = 7
    
    schools = School.objects.all()
    for school in schools:
        school_students = students.filter(school=school).count()
        ws_summary[f'A{row}'] = school.name
        ws_summary[f'B{row}'] = school_students
        row += 1
    
    # Department-wise count
    ws_summary[f'A{row+1}'] = "Department-wise Distribution:"
    ws_summary[f'A{row+1}'].font = Font(bold=True)
    row += 2
    
    departments = Department.objects.all()
    for dept in departments:
        dept_students = students.filter(department=dept).count()
        ws_summary[f'A{row}'] = f"{dept.name} ({dept.school.name})"
        ws_summary[f'B{row}'] = dept_students
        row += 1
    
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Registered_Students_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@staff_member_required
def export_combined_course_report(request, course_id):
    """Export combined report for all teachers of a specific course"""
    
    course = get_object_or_404(Course, id=course_id)
    
    # Get all forms for this course
    forms = FeedbackForm.objects.filter(course=course).select_related(
        'teacher', 'course__department', 'course__department__school'
    ).order_by('teacher__name')
    
    if not forms.exists():
        # Return empty response with message
        response = HttpResponse("No feedback forms found for this course")
        return response
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styling
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================
    # SHEET 1: SUMMARY
    # ============================================
    
    ws_summary = wb.create_sheet("Course Summary")
    
    # Title
    ws_summary['A1'] = "Combined Feedback Report"
    ws_summary['A1'].font = Font(bold=True, size=16, color="6366F1")
    
    # Course Information
    ws_summary['A3'] = "Course Code:"
    ws_summary['B3'] = course.code
    ws_summary['A4'] = "Course Name:"
    ws_summary['B4'] = course.name
    ws_summary['A5'] = "Department:"
    ws_summary['B5'] = course.department.name
    ws_summary['A6'] = "School:"
    ws_summary['B6'] = course.department.school.name
    ws_summary['A7'] = "Semester:"
    ws_summary['B7'] = f"{course.semester} / Year {course.year}"
    ws_summary['A8'] = "Total Teachers:"
    ws_summary['B8'] = forms.count()
    ws_summary['A9'] = "Generated On:"
    ws_summary['B9'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Make labels bold
    for row in range(3, 10):
        ws_summary[f'A{row}'].font = Font(bold=True)
    
    # Teacher-wise Summary Table
    ws_summary['A11'] = "Teacher-wise Summary"
    ws_summary['A11'].font = Font(bold=True, size=14, color="6366F1")
    
    # Headers
    headers = ['S.No', 'Teacher Name', 'Employee ID', 'Total Submissions', 'Form Status']
    for col_num, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=12, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    row_num = 13
    for idx, form in enumerate(forms, 1):
        submission_count = form.submissions.count()
        
        ws_summary.cell(row=row_num, column=1, value=idx)
        ws_summary.cell(row=row_num, column=2, value=form.teacher.name)
        ws_summary.cell(row=row_num, column=3, value=form.teacher.employee_id or 'N/A')
        ws_summary.cell(row=row_num, column=4, value=submission_count)
        ws_summary.cell(row=row_num, column=5, value='Active' if form.is_active else 'Inactive')
        
        # Apply borders
        for col_num in range(1, 6):
            ws_summary.cell(row=row_num, column=col_num).border = border
        
        row_num += 1
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 8
    ws_summary.column_dimensions['B'].width = 30
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 20
    ws_summary.column_dimensions['E'].width = 15
    
    # ============================================
    # SHEET 2-N: EACH TEACHER'S DETAILED REPORT
    # ============================================
    
    for form in forms:
        # Create sheet for each teacher
        teacher_name = form.teacher.name[:25]  # Limit sheet name length
        sheet_name = f"{teacher_name}"
        
        # Make sheet name unique
        if sheet_name in wb.sheetnames:
            sheet_name = f"{teacher_name}_{form.id}"
        
        ws = wb.create_sheet(sheet_name)
        
        # Teacher Header
        ws['A1'] = "Teacher Feedback Report"
        ws['A1'].font = Font(bold=True, size=14, color="6366F1")
        
        ws['A3'] = "Teacher Name:"
        ws['B3'] = form.teacher.name
        ws['A4'] = "Employee ID:"
        ws['B4'] = form.teacher.employee_id or 'N/A'
        ws['A5'] = "Course:"
        ws['B5'] = f"{course.code} - {course.name}"
        ws['A6'] = "Total Submissions:"
        ws['B6'] = form.submissions.count()
        
        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)
        
        current_row = 8
        
        # Get questions
        questions = form.questions.all().prefetch_related('options').order_by('order')
        
        for question in questions:
            # Question header
            ws[f'A{current_row}'] = f"Q{question.order}: {question.question_text}"
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 1
            
            if question.question_type == 'mcq':
                # MCQ Results
                option_counts = Response.objects.filter(
                    question=question,
                    mcq_answer__isnull=False
                ).values(
                    'mcq_answer__option_text',
                    'mcq_answer__order'
                ).annotate(
                    count=Count('id')
                ).order_by('mcq_answer__order')
                
                total_responses = sum(item['count'] for item in option_counts)
                
                # Table headers
                ws[f'A{current_row}'] = "Option"
                ws[f'B{current_row}'] = "Count"
                ws[f'C{current_row}'] = "Percentage"
                ws[f'D{current_row}'] = "Visual"
                
                for col in ['A', 'B', 'C', 'D']:
                    cell = ws[f'{col}{current_row}']
                    cell.fill = subheader_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                current_row += 1
                
                # Data rows
                for item in option_counts:
                    ws[f'A{current_row}'] = item['mcq_answer__option_text']
                    ws[f'B{current_row}'] = item['count']
                    
                    if total_responses > 0:
                        percentage = (item['count'] / total_responses) * 100
                        ws[f'C{current_row}'] = f"{percentage:.1f}%"
                        
                        # Visual bar (using repeated characters)
                        bar_length = int(percentage / 5)  # Scale to fit
                        ws[f'D{current_row}'] = '█' * bar_length
                    else:
                        ws[f'C{current_row}'] = "0%"
                        ws[f'D{current_row}'] = ''
                    
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{current_row}'].border = border
                    
                    current_row += 1
                
                current_row += 1  # Empty row
                
            elif question.question_type == 'text':
                # Text Responses
                text_responses = Response.objects.filter(
                    question=question,
                    text_answer__isnull=False
                ).exclude(
                    text_answer=''
                ).select_related(
                    'submission__student'
                ).order_by('submission__submitted_at')
                
                if text_responses.exists():
                    # Table headers
                    ws[f'A{current_row}'] = "Student"
                    ws[f'B{current_row}'] = "Response"
                    ws[f'C{current_row}'] = "Date"
                    
                    for col in ['A', 'B', 'C']:
                        cell = ws[f'{col}{current_row}']
                        cell.fill = subheader_fill
                        cell.font = header_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                    
                    current_row += 1
                    
                    # Data rows
                    for response in text_responses:
                        ws[f'A{current_row}'] = response.submission.student.name
                        ws[f'B{current_row}'] = response.text_answer
                        ws[f'C{current_row}'] = response.submission.submitted_at.strftime("%Y-%m-%d")
                        
                        # Wrap text
                        ws[f'B{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
                        
                        for col in ['A', 'B', 'C']:
                            ws[f'{col}{current_row}'].border = border
                        
                        current_row += 1
                else:
                    ws[f'A{current_row}'] = "No text responses"
                    current_row += 1
                
                current_row += 1  # Empty row
        
        # Adjust column widths for teacher sheet
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
    
    # ============================================
    # SHEET: COMPARATIVE ANALYSIS
    # ============================================
    
    ws_compare = wb.create_sheet("Comparative Analysis")
    
    # Header with Course Info
    ws_compare['A1'] = "Comparative Analysis Report"
    ws_compare['A1'].font = Font(bold=True, size=16, color="6366F1")
    ws_compare.merge_cells('A1:E1')
    
    # Course Information in Comparative Sheet
    ws_compare['A3'] = "Course Code:"
    ws_compare['B3'] = course.code
    ws_compare['A4'] = "Course Name:"
    ws_compare['B4'] = course.name
    ws_compare['A5'] = "Department:"
    ws_compare['B5'] = course.department.name
    ws_compare['A6'] = "School:"
    ws_compare['B6'] = course.department.school.name
    ws_compare['A7'] = "Semester / Year:"
    ws_compare['B7'] = f"Semester {course.semester} / Year {course.year}"
    ws_compare['A8'] = "Number of Teachers:"
    ws_compare['B8'] = forms.count()
    
    # Make labels bold
    for row in range(3, 9):
        ws_compare[f'A{row}'].font = Font(bold=True)
        ws_compare[f'A{row}'].fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    
    ws_compare['A10'] = "This analysis compares feedback responses across all teachers for this course"
    ws_compare['A10'].font = Font(italic=True, color="64748B")
    ws_compare.merge_cells('A10:E10')
    
    # Find common questions (assuming first form has template questions)
    if forms.exists():
        first_form = forms.first()
        common_questions = first_form.questions.filter(question_type='mcq').order_by('order')
        
        current_row = 12
        
        for question in common_questions:
            # Question Header with highlighting
            ws_compare[f'A{current_row}'] = f"Q{question.order}: {question.question_text}"
            ws_compare[f'A{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws_compare[f'A{current_row}'].fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
            
            # Merge across all teacher columns
            last_col = openpyxl.utils.get_column_letter(1 + forms.count())
            ws_compare.merge_cells(f'A{current_row}:{last_col}{current_row}')
            current_row += 1
            
            # Headers: Option + Teacher names
            ws_compare[f'A{current_row}'] = "Response Options"
            ws_compare[f'A{current_row}'].fill = subheader_fill
            ws_compare[f'A{current_row}'].font = header_font
            ws_compare[f'A{current_row}'].border = border
            ws_compare[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            col_idx = 2
            for form in forms:
                cell = ws_compare.cell(row=current_row, column=col_idx)
                # Show teacher name with course name
                teacher_label = f"{form.teacher.name}\n{course.code} - {course.name}"
                cell.value = teacher_label
                cell.fill = subheader_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                col_idx += 1
            
            current_row += 1
            
            # Get options from first question
            options = question.options.all().order_by('order')
            
            for option in options:
                ws_compare[f'A{current_row}'] = option.option_text
                ws_compare[f'A{current_row}'].border = border
                ws_compare[f'A{current_row}'].font = Font(bold=True)
                ws_compare[f'A{current_row}'].fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                
                col_idx = 2
                for form in forms:
                    # Find equivalent question in this form
                    try:
                        form_question = form.questions.get(order=question.order, question_type='mcq')
                        form_option = form_question.options.get(order=option.order)
                        
                        count = Response.objects.filter(
                            question=form_question,
                            mcq_answer=form_option
                        ).count()
                        
                        total = Response.objects.filter(question=form_question).count()
                        
                        if total > 0:
                            percentage = (count / total) * 100
                            cell_value = f"{count} ({percentage:.0f}%)"
                        else:
                            cell_value = "0 (0%)"
                        
                        cell = ws_compare.cell(row=current_row, column=col_idx)
                        cell.value = cell_value
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                        
                        # Color coding based on count (optional - for visual appeal)
                        if percentage >= 50:
                            cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                        elif percentage >= 30:
                            cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        
                    except:
                        cell = ws_compare.cell(row=current_row, column=col_idx)
                        cell.value = "N/A"
                        cell.border = border
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                    
                    col_idx += 1
                
                current_row += 1
            
            # Add average row
            ws_compare[f'A{current_row}'] = "Total Responses"
            ws_compare[f'A{current_row}'].font = Font(bold=True, italic=True)
            ws_compare[f'A{current_row}'].fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            ws_compare[f'A{current_row}'].border = border
            
            col_idx = 2
            for form in forms:
                try:
                    form_question = form.questions.get(order=question.order, question_type='mcq')
                    total = Response.objects.filter(question=form_question).count()
                    
                    cell = ws_compare.cell(row=current_row, column=col_idx)
                    cell.value = total
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
                except:
                    cell = ws_compare.cell(row=current_row, column=col_idx)
                    cell.value = "N/A"
                    cell.border = border
                
                col_idx += 1
            
            current_row += 3  # Extra space between questions
    
    # Adjust column widths
    ws_compare.column_dimensions['A'].width = 40
    for col_idx in range(2, 2 + forms.count()):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws_compare.column_dimensions[col_letter].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Combined_Report_{course.code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@staff_member_required
def export_students_list(request):
    """Export all registered students to Excel file"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registered Students"
    
    # Styling
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['S.No', 'Roll Number', 'Name', 'School', 'Department', 'Enrolled Courses', 'Registration Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get all students with related data
    students = Student.objects.filter(is_staff=False).select_related(
        'school', 'department'
    ).prefetch_related('enrolled_courses__course').order_by('school', 'department', 'name')
    
    # Write data
    row_num = 2
    for idx, student in enumerate(students, 1):
        # Get enrolled courses
        courses = student.enrolled_courses.all()
        course_list = ', '.join([f"{sc.course.code}" for sc in courses]) if courses.exists() else 'None'
        
        # Write row
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=student.roll_number)
        ws.cell(row=row_num, column=3, value=student.name)
        ws.cell(row=row_num, column=4, value=student.school.name if student.school else 'N/A')
        ws.cell(row=row_num, column=5, value=student.department.name if student.department else 'N/A')
        ws.cell(row=row_num, column=6, value=course_list)
        ws.cell(row=row_num, column=7, value=student.date_joined.strftime("%Y-%m-%d"))
        
        # Apply borders
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).border = border
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 18
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "Student Registration Summary"
    ws_summary['A1'].font = Font(bold=True, size=16, color="6366F1")
    
    ws_summary['A3'] = "Total Students:"
    ws_summary['B3'] = students.count()
    ws_summary['A4'] = "Generated On:"
    ws_summary['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # School-wise count
    ws_summary['A6'] = "School-wise Distribution:"
    ws_summary['A6'].font = Font(bold=True)
    row = 7
    
    schools = School.objects.all()
    for school in schools:
        school_students = students.filter(school=school).count()
        ws_summary[f'A{row}'] = school.name
        ws_summary[f'B{row}'] = school_students
        row += 1
    
    # Department-wise count
    ws_summary[f'A{row+1}'] = "Department-wise Distribution:"
    ws_summary[f'A{row+1}'].font = Font(bold=True)
    row += 2
    
    departments = Department.objects.all()
    for dept in departments:
        dept_students = students.filter(department=dept).count()
        ws_summary[f'A{row}'] = f"{dept.name} ({dept.school.name})"
        ws_summary[f'B{row}'] = dept_students
        row += 1
    
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Registered_Students_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response